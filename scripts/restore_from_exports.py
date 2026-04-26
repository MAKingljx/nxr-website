#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import shutil
import sqlite3
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "Data"


TEMP_EXPORT_COLUMNS = [
    "cert_id",
    "card_name",
    "year",
    "brand",
    "variety",
    "pop",
    "language",
    "set_name",
    "card_number",
    "centering",
    "edges",
    "corners",
    "surface",
    "final_grade",
    "final_grade_text",
    "front_image",
    "back_image",
    "entry_notes",
    "entry_by",
    "entry_date",
    "status",
    "created_at",
    "updated_at",
    "upload_status",
    "upload_started",
    "upload_completed",
    "upload_error",
    "server_response",
    "published_front_image",
    "published_back_image",
    "approved_at",
    "approval_sequence",
]


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    return text


def normalize_numeric(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_int(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None


def row_score(row):
    score = 0
    for key, value in row.items():
        if key == "_sources":
            continue
        if value not in (None, ""):
            score += 1
    return score


def load_export_rows(exports_dir: Path):
    merged = {}
    file_rows = {}
    for path in sorted(exports_dir.glob("*.xlsx")):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [normalize_text(cell) for cell in next(rows)]
        source_count = 0
        for values in rows:
            raw = dict(zip(headers, values))
            cert_id = normalize_text(raw.get("cert_id"))
            if not cert_id:
                continue
            source_count += 1
            normalized = {
                "cert_id": cert_id,
                "card_name": normalize_text(raw.get("card_name")),
                "year": normalize_text(raw.get("year")),
                "brand": normalize_text(raw.get("brand")),
                "variety": normalize_text(raw.get("variety")),
                "pop": normalize_text(raw.get("pop")) or "1",
                "language": normalize_text(raw.get("language")) or "EN",
                "set_name": normalize_text(raw.get("set_name")),
                "card_number": normalize_text(raw.get("card_number")),
                "centering": normalize_numeric(raw.get("centering")),
                "edges": normalize_numeric(raw.get("edges")),
                "corners": normalize_numeric(raw.get("corners")),
                "surface": normalize_numeric(raw.get("surface")),
                "final_grade": normalize_numeric(raw.get("final_grade")),
                "final_grade_text": normalize_text(raw.get("final_grade_text")),
                "front_image": normalize_text(raw.get("front_image")),
                "back_image": normalize_text(raw.get("back_image")),
                "entry_notes": normalize_text(raw.get("entry_notes")),
                "entry_by": normalize_text(raw.get("entry_by")) or "admin",
                "entry_date": normalize_text(raw.get("entry_date")),
                "status": normalize_text(raw.get("status")) or "approved",
                "created_at": normalize_text(raw.get("created_at")),
                "updated_at": normalize_text(raw.get("updated_at")),
                "upload_status": normalize_text(raw.get("upload_status")) or "not_started",
                "upload_started": normalize_text(raw.get("upload_started")),
                "upload_completed": normalize_text(raw.get("upload_completed")),
                "upload_error": normalize_text(raw.get("upload_error")),
                "server_response": normalize_text(raw.get("server_response")),
                "published_front_image": normalize_text(raw.get("published_front_image")),
                "published_back_image": normalize_text(raw.get("published_back_image")),
                "approved_at": normalize_text(raw.get("approved_at")),
                "approval_sequence": normalize_int(raw.get("approval_sequence")),
                "_sources": [path.name],
            }
            existing = merged.get(cert_id)
            if existing is None:
                merged[cert_id] = normalized
                continue
            if row_score(normalized) > row_score(existing):
                winner, loser = normalized, existing
            else:
                winner, loser = existing, normalized
            for key, value in loser.items():
                if key == "_sources":
                    continue
                if winner.get(key) in (None, "") and value not in (None, ""):
                    winner[key] = value
            winner["_sources"] = sorted(set(existing["_sources"] + normalized["_sources"]))
            merged[cert_id] = winner
        file_rows[path.name] = source_count
    return merged, file_rows


def ensure_parent(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)


def copy_seed_db(source: Path, destination: Path):
    ensure_parent(destination)
    shutil.copy2(source, destination)


def clear_table(conn: sqlite3.Connection, table: str):
    conn.execute(f"DELETE FROM {table}")
    conn.commit()


def get_existing_cert_ids(conn: sqlite3.Connection, table: str):
    return {row[0] for row in conn.execute(f"SELECT cert_id FROM {table}")}


def restore_temp_cards(conn: sqlite3.Connection, rows_by_cert):
    existing = get_existing_cert_ids(conn, "temp_cards")
    inserted = 0
    updated = 0
    status_counter = Counter()
    for cert_id, row in sorted(rows_by_cert.items()):
        values = []
        for column in TEMP_EXPORT_COLUMNS:
            value = row.get(column)
            if column in {"centering", "edges", "corners", "surface", "final_grade"}:
                value = 1 if value is None else value
            elif column == "approval_sequence":
                value = row.get(column)
            elif value is None:
                value = ""
            values.append(value)

        if cert_id in existing:
            set_clause = ", ".join(f"{column} = ?" for column in TEMP_EXPORT_COLUMNS if column != "cert_id")
            update_values = [row.get(column) if column != "approval_sequence" else row.get(column) for column in TEMP_EXPORT_COLUMNS if column != "cert_id"]
            normalized_updates = []
            for column, value in zip([c for c in TEMP_EXPORT_COLUMNS if c != "cert_id"], update_values):
                if column in {"centering", "edges", "corners", "surface", "final_grade"}:
                    normalized_updates.append(1 if value is None else value)
                elif column == "approval_sequence":
                    normalized_updates.append(value)
                elif value is None:
                    normalized_updates.append("")
                else:
                    normalized_updates.append(value)
            conn.execute(
                f"UPDATE temp_cards SET {set_clause} WHERE cert_id = ?",
                normalized_updates + [cert_id],
            )
            updated += 1
        else:
            placeholders = ", ".join("?" for _ in TEMP_EXPORT_COLUMNS)
            conn.execute(
                f"INSERT INTO temp_cards ({', '.join(TEMP_EXPORT_COLUMNS)}) VALUES ({placeholders})",
                values,
            )
            inserted += 1
        status_counter[row.get("status") or ""] += 1
    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM temp_cards").fetchone()[0]
    return {
        "inserted": inserted,
        "updated": updated,
        "total_after_restore": total,
        "status_counts_from_exports": dict(status_counter),
    }


def build_cards_payload(row):
    cert_id = row["cert_id"]
    front_image = row.get("published_front_image") or ""
    back_image = row.get("published_back_image") or ""

    if not front_image or not back_image:
        server_response = row.get("server_response") or ""
        if server_response:
            try:
                payload = json.loads(server_response)
            except json.JSONDecodeError:
                payload = {}
            front_image = front_image or normalize_text(payload.get("front_image"))
            back_image = back_image or normalize_text(payload.get("back_image"))

    image = front_image or ""
    created_at = row.get("created_at") or row.get("entry_date") or ""
    updated_at = row.get("updated_at") or created_at

    return {
        "cert_id": cert_id,
        "card_name": row.get("card_name") or "",
        "grade": row.get("final_grade_text") or "",
        "year": row.get("year") or "",
        "brand": row.get("brand") or "",
        "player": "",
        "variety": row.get("variety") or "",
        "image": image,
        "pop": row.get("pop") or "1",
        "back_image": back_image,
        "front_image": front_image,
        "qr_url": f"/card/{cert_id}",
        "centering": 0 if row.get("centering") is None else row["centering"],
        "edges": 0 if row.get("edges") is None else row["edges"],
        "corners": 0 if row.get("corners") is None else row["corners"],
        "surface": 0 if row.get("surface") is None else row["surface"],
        "language": row.get("language") or "EN",
        "set_name": row.get("set_name") or "",
        "card_number": row.get("card_number") or "",
        "grading_phase": "human_only",
        "data_version": 1,
        "created_at": created_at,
        "updated_at": updated_at,
        "ai_model_version": "",
        "ai_confidence": 0,
        "ai_grade": None,
        "ai_centering": None,
        "ai_edges": None,
        "ai_corners": None,
        "ai_surface": None,
        "final_grade": 0 if row.get("final_grade") is None else row["final_grade"],
        "decision_method": "human_only",
        "decision_notes": row.get("entry_notes") or "",
        "ai_front_analysis": "",
        "ai_back_analysis": "",
        "has_ai_analysis": 0,
        "final_grade_text": row.get("final_grade_text") or "",
    }


def restore_cards(conn: sqlite3.Connection, rows_by_cert):
    existing = get_existing_cert_ids(conn, "cards")
    inserted = 0
    updated = 0
    with_images = 0
    for cert_id, row in sorted(rows_by_cert.items()):
        payload = build_cards_payload(row)
        if payload["front_image"] and payload["back_image"]:
            with_images += 1
        columns = list(payload.keys())
        if cert_id in existing:
            set_clause = ", ".join(f"{column} = ?" for column in columns if column != "cert_id")
            conn.execute(
                f"UPDATE cards SET {set_clause} WHERE cert_id = ?",
                [payload[column] for column in columns if column != "cert_id"] + [cert_id],
            )
            updated += 1
        else:
            placeholders = ", ".join("?" for _ in columns)
            conn.execute(
                f"INSERT INTO cards ({', '.join(columns)}) VALUES ({placeholders})",
                [payload[column] for column in columns],
            )
            inserted += 1
    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM cards").fetchone()[0]
    return {
        "inserted": inserted,
        "updated": updated,
        "total_after_restore": total,
        "rows_with_published_images": with_images,
    }


def parse_args():
    parser = argparse.ArgumentParser(description="Restore local SQLite databases from approved card Excel exports.")
    parser.add_argument("--exports-dir", required=True, help="Directory containing exported xlsx files")
    parser.add_argument("--output-dir", required=True, help="Directory to store restored database copies and report")
    parser.add_argument(
        "--truncate-existing",
        action="store_true",
        help="Clear existing rows from the copied seed databases before import",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    exports_dir = Path(args.exports_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    rows_by_cert, file_rows = load_export_rows(exports_dir)

    cards_out = output_dir / "cards_restored.db"
    temp_out = output_dir / "temp_cards_restored.db"
    copy_seed_db(DATA_DIR / "cards.db", cards_out)
    copy_seed_db(DATA_DIR / "temp_cards.db", temp_out)

    conn_cards = sqlite3.connect(cards_out)
    conn_temp = sqlite3.connect(temp_out)
    try:
        if args.truncate_existing:
            clear_table(conn_cards, "cards")
            clear_table(conn_temp, "temp_cards")
        cards_summary = restore_cards(conn_cards, rows_by_cert)
        temp_summary = restore_temp_cards(conn_temp, rows_by_cert)
    finally:
        conn_cards.close()
        conn_temp.close()

    summary = {
        "exports_dir": str(exports_dir),
        "xlsx_files": len(file_rows),
        "rows_per_file": file_rows,
        "unique_cert_ids_restored": len(rows_by_cert),
        "cards": cards_summary,
        "temp_cards": temp_summary,
        "output": {
            "cards_db": str(cards_out),
            "temp_cards_db": str(temp_out),
        },
        "truncate_existing": args.truncate_existing,
    }
    summary_path = output_dir / "restore_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
